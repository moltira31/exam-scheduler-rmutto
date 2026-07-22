import datetime
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

# ==================== 1. ตั้งค่าหน้าจอโปรแกรม ====================
st.set_page_config(
    page_title="ระบบจัดตารางสอบ - มทร.ตะวันออก วิทยาเขตจันทบุรี",
    page_icon="🏫",
    layout="wide",
)

USER_CREDENTIALS = {"admin1": "rmutto123", "registry_staff": "rmutto456"}

# --- Session State สำหรับล็อกอินและจดจำตารางสอบเดิม ---
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "username" not in st.session_state:
    st.session_state["username"] = ""

# ความจำตารางสอบสะสม (Memory across multiple uploads in same semester)
if "history_schedule" not in st.session_state:
    st.session_state["history_schedule"] = pd.DataFrame()


def logout():
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""
    st.rerun()


def reset_semester_data():
    st.session_state["history_schedule"] = pd.DataFrame()
    st.success("ล้างข้อมูลตารางสอบเดิมในระบบเรียบร้อยแล้ว! พร้อมสำหรับการจัดสอบเทอมใหม่")


# ==================== 2. ฟังก์ชันสร้างช่วงเวลาสอบ ====================
def generate_time_slots(start_date, end_date, daily_slots):
    """สร้างรายการวันและช่วงเวลาสอบทั้งหมดที่เป็นไปได้ (เว้นวันเสาร์-อาทิตย์)"""
    slots = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:  # จันทร์ - ศุกร์
            date_str = current_date.strftime("%d/%m/%Y")
            for slot_name in daily_slots:
                slots.append(f"{date_str} ({slot_name})")
        current_date += datetime.timedelta(days=1)
    return slots


# ==================== 3. อัลกอริทึมจัดตารางสอบ (เช็คซ้อนกับรอบก่อนหน้า) ====================
def auto_schedule_exams(
    df_subjects,
    df_rooms,
    slots_m,
    slots_f,
    existing_schedule=None,
    shuffle_invigilators=False,
    avoid_own_dept=True,
):
    """
    จัดตารางสอบโดยตรวจสอบเงื่อนไขไม่ให้เกิดข้อผิดพลาด:
    - ดึงตารางสอบรอบก่อนหน้า (existing_schedule) มาจองพื้นที่ล่วงหน้าเพื่อไม่ให้ซ้อน
    1. ห้ามกลุ่มนักศึกษาซ้อนเวลากัน
    2. ห้ามใช้ห้องสอบซ้อนกัน และความจุต้องพอ
    3. ห้ามอาจารย์คุมสอบติดภาระงานซ้อนกัน
    """
    student_group_occupancy = {}  # {slot: set(กลุ่มเรียน)}
    room_occupancy = {}  # {slot: set(รหัสห้อง)}
    invigilator_occupancy = {}  # {slot: set(ชื่อผู้คุมสอบ)}

    # --- โหลดข้อมูลสอบเดิมเข้าระบบ (ถ้ามี) เพื่อป้องกันการจัดซ้อน ---
    if existing_schedule is not None and not existing_schedule.empty:
        for _, ex_row in existing_schedule.iterrows():
            grp = str(ex_row.get("กลุ่มเรียน", "")).strip()

            # สกัด Slot และ ห้องสอบ กลางภาค
            m_info = str(ex_row.get("วันเวลาสอบ_M", ""))
            if m_info and " [ห้อง " in m_info:
                m_slot = m_info.split(" [ห้อง ")[0].strip()
                m_rm = m_info.split(" [ห้อง ")[1].replace("]", "").strip()
                m_inv = str(ex_row.get("ผู้คุมสอบ_M", "")).strip()

                student_group_occupancy.setdefault(m_slot, set()).add(grp)
                if m_rm:
                    room_occupancy.setdefault(m_slot, set()).add(m_rm)
                if m_inv:
                    invigilator_occupancy.setdefault(m_slot, set()).add(m_inv)

            # สกัด Slot และ ห้องสอบ ปลายภาค
            f_info = str(ex_row.get("วันเวลาสอบ_F", ""))
            if f_info and " [ห้อง " in f_info:
                f_slot = f_info.split(" [ห้อง ")[0].strip()
                f_rm = f_info.split(" [ห้อง ")[1].replace("]", "").strip()
                f_inv = str(ex_row.get("ผู้คุมสอบ_F", "")).strip()

                student_group_occupancy.setdefault(f_slot, set()).add(grp)
                if f_rm:
                    room_occupancy.setdefault(f_slot, set()).add(f_rm)
                if f_inv:
                    invigilator_occupancy.setdefault(f_slot, set()).add(f_inv)

    all_teachers = list(df_subjects["ชื่อผู้สอน"].dropna().unique())
    results = []

    for idx, row in df_subjects.iterrows():
        subj_code = row.get("รหัสวิชา", f"SUBJ-{idx}")
        subj_name = row.get("ชื่อวิชา", "")
        instructor = row.get("ชื่อผู้สอน", "อาจารย์ผู้สอน")
        faculty = row.get("คณะ", row.get("สังกัดคณะ", "คณะเทคโนโลยีอุตสาหกรรมการเกษตร"))
        dept = row.get("สังกัดสาขา", "")
        group = str(row.get("กลุ่มเรียน", "")).strip()
        students_count = int(row.get("จำนวนผู้เข้าสอบ", 0))
        exam_type = row.get("ประเภทการสอบ", "ทฤษฎี")

        valid_rooms = df_rooms[
            (df_rooms["ความจุสอบ"] >= students_count)
            & (df_rooms["สถานะ"] == "ใช้งานได้")
        ]
        if exam_type == "ปฏิบัติคอมพิวเตอร์":
            valid_rooms = valid_rooms[
                valid_rooms["ประเภท"] == "ห้องปฏิบัติการคอมพิวเตอร์"
            ]

        def get_available_invigilator(slot):
            used_teachers = invigilator_occupancy.get(slot, set())
            if not shuffle_invigilators:
                return (
                    instructor if instructor not in used_teachers else None
                )

            for t in all_teachers:
                if t not in used_teachers:
                    if avoid_own_dept and len(all_teachers) > 1:
                        if t != instructor:
                            return t
                    else:
                        return t
            return None

        # --- จัดสอบกลางภาค (Midterm - M) ---
        m_slot_selected = ""
        m_room_selected = ""
        m_invigilator = ""

        hrs_m = row.get("ชั่วโมงสอบ_M", row.get("ชั่วโมงสอบ (นาที)", 90))
        if pd.notna(hrs_m) and str(hrs_m).strip() != "" and float(hrs_m) > 0:
            for slot in slots_m:
                if group in student_group_occupancy.get(slot, set()):
                    continue

                invig = get_available_invigilator(slot)
                if not invig:
                    continue

                used_rooms = room_occupancy.get(slot, set())
                avail_room = None
                for _, r in valid_rooms.iterrows():
                    if r["รหัสห้อง"] not in used_rooms:
                        avail_room = r["รหัสห้อง"]
                        break

                if avail_room:
                    m_slot_selected = slot
                    m_room_selected = avail_room
                    m_invigilator = invig

                    student_group_occupancy.setdefault(slot, set()).add(group)
                    invigilator_occupancy.setdefault(slot, set()).add(invig)
                    room_occupancy.setdefault(slot, set()).add(avail_room)
                    break

        # --- จัดสอบปลายภาค (Final - F) ---
        f_slot_selected = ""
        f_room_selected = ""
        f_invigilator = ""

        hrs_f = row.get("ชั่วโมงสอบ_F", row.get("ชั่วโมงสอบ (นาที)", 120))
        if pd.notna(hrs_f) and str(hrs_f).strip() != "" and float(hrs_f) > 0:
            for slot in slots_f:
                if group in student_group_occupancy.get(slot, set()):
                    continue

                invig = get_available_invigilator(slot)
                if not invig:
                    continue

                used_rooms = room_occupancy.get(slot, set())
                avail_room = None
                for _, r in valid_rooms.iterrows():
                    if r["รหัสห้อง"] not in used_rooms:
                        avail_room = r["รหัสห้อง"]
                        break

                if avail_room:
                    f_slot_selected = slot
                    f_room_selected = avail_room
                    f_invigilator = invig

                    student_group_occupancy.setdefault(slot, set()).add(group)
                    invigilator_occupancy.setdefault(slot, set()).add(invig)
                    room_occupancy.setdefault(slot, set()).add(avail_room)
                    break

        # ปรับการแสดงผลชั่วโมงสอบ
        val_hrs_m = float(hrs_m)/60 if float(hrs_m) > 10 else float(hrs_m) if hrs_m else ""
        val_hrs_f = float(hrs_f)/60 if float(hrs_f) > 10 else float(hrs_f) if hrs_f else 2

        results.append({
            "คณะ": faculty,
            "รหัสวิชา": subj_code,
            "ชื่อวิชา": subj_name,
            "ชื่อผู้สอน": instructor,
            "สังกัดสาขา": dept,
            "กลุ่มเรียน": group,
            "จำนวนผู้เข้าสอบ": students_count,
            "ชั่วโมงสอบ_M": val_hrs_m if m_slot_selected else "",
            "ชั่วโมงสอบ_F": val_hrs_f if f_slot_selected else "",
            "วันเวลาสอบ_M": (
                f"{m_slot_selected} [ห้อง {m_room_selected}]"
                if m_slot_selected
                else "ไม่มีสอบ"
            ),
            "วันเวลาสอบ_F": (
                f"{f_slot_selected} [ห้อง {f_room_selected}]"
                if f_slot_selected
                else "ไม่มีสอบ"
            ),
            "ผู้คุมสอบ_M": m_invigilator,
            "ผู้คุมสอบ_F": f_invigilator,
        })

    return pd.DataFrame(results)


# ==================== 4. ฟังก์ชันสร้างไฟล์ Excel (แยกหลายคณะตามจริง) ====================
def generate_excel_report(df_schedule, title_m, title_f):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายวิชาที่จัดสอบในตารางMF"
    ws.views.sheetView[0].showGridLines = True

    HEADER_BG = "BDD7EE"
    FACULTY_BG = "FCE4D6"
    GRAY_FILL = "D9D9D9"

    font_title = Font(name="TH SarabunPSK", size=16, bold=True)
    font_subtitle = Font(name="TH SarabunPSK", size=14, bold=True)
    font_header = Font(name="TH SarabunPSK", size=12, bold=True)
    font_faculty = Font(
        name="TH SarabunPSK", size=13, bold=True, color="C00000"
    )
    font_data = Font(name="TH SarabunPSK", size=11)
    font_green_m = Font(name="TH SarabunPSK", size=11, color="548235", bold=True)
    font_red_f = Font(name="TH SarabunPSK", size=11, color="C00000", bold=True)

    align_center = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_side = Side(style="thin", color="000000")
    border_all = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # หัวเรื่องใหญ่
    ws.merge_cells("A1:K1")
    ws["A1"] = "รายวิชาที่สอบในตาราง 1/2569"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:K2")
    ws["A2"] = f"สอบกลางภาค {title_m} และ สอบปลายภาค {title_f}"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_center

    # หัวคอลัมน์
    headers = [
        ("A3", "ที่"),
        ("B3", "รหัสวิชา"),
        ("C3", "รายวิชา"),
        ("D3", "ผู้สอน"),
        ("E3", "สำรองที่นั่ง"),
        ("F3", "กลุ่ม"),
        ("G3", "ลง"),
        ("H3", "M/F"),
        ("I3", "ชั่วโมง\nสอบ"),
        ("J3", "วัน/เวลาสอบ"),
        ("K3", "ผู้คุมสอบ"),
    ]

    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = PatternFill(
            start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid"
        )
        cell.border = border_all

    current_row = 4
    global_item_no = 1

    # กลุ่มข้อมูลตามคณะที่มีจริงใน DataFrame
    faculties = df_schedule["คณะ"].unique() if "คณะ" in df_schedule.columns else ["คณะเทคโนโลยีอุตสาหกรรมการเกษตร"]

    for fac in faculties:
        df_fac = df_schedule[df_schedule["คณะ"] == fac]
        
        # นับจำนวนวิชาที่จัดสอบ M และ F ของคณะนั้นๆ
        m_count = len(df_fac[df_fac["วันเวลาสอบ_M"] != "ไม่มีสอบ"])
        f_count = len(df_fac[df_fac["วันเวลาสอบ_F"] != "ไม่มีสอบ"])

        # หัวข้อคณะ (สีกระทิง/ส้มอ่อน ตามแบบ)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        fac_cell = ws.cell(row=current_row, column=1)
        fac_cell.value = f"{fac} (M={m_count} , F={f_count})"
        fac_cell.font = font_faculty
        fac_cell.alignment = align_center
        fac_cell.fill = PatternFill(start_color=FACULTY_BG, end_color=FACULTY_BG, fill_type="solid")
        
        for c in range(1, 12):
            ws.cell(row=current_row, column=c).border = border_all
        
        current_row += 1

        # แสดงรายการวิชาในคณะนั้น
        for idx, row in df_fac.iterrows():
            r_m = current_row
            r_f = current_row + 1

            for col_idx in [1, 2, 3, 4, 6, 7]:
                ws.merge_cells(start_row=r_m, end_row=r_f, start_column=col_idx, end_column=col_idx)

            ws.cell(row=r_m, column=1, value=global_item_no).alignment = align_center
            ws.cell(row=r_m, column=2, value=row.get("รหัสวิชา", "")).alignment = align_center
            ws.cell(row=r_m, column=3, value=row.get("ชื่อวิชา", "")).alignment = align_left
            ws.cell(row=r_m, column=4, value=row.get("ชื่อผู้สอน", "")).alignment = align_center
            ws.cell(row=r_m, column=6, value=row.get("กลุ่มเรียน", "")).alignment = align_center
            ws.cell(row=r_m, column=7, value=row.get("จำนวนผู้เข้าสอบ", "")).alignment = align_center

            # M Row
            ws.cell(row=r_m, column=8, value="M").font = font_green_m
            ws.cell(row=r_m, column=8).alignment = align_center
            hrs_m = row.get("ชั่วโมงสอบ_M", "")
            ws.cell(row=r_m, column=9, value=hrs_m).alignment = align_center
            ws.cell(row=r_m, column=10, value=row.get("วันเวลาสอบ_M", "")).alignment = align_center
            ws.cell(row=r_m, column=11, value=row.get("ผู้คุมสอบ_M", "")).alignment = align_center

            if not hrs_m or hrs_m == "":
                ws.cell(row=r_m, column=9).fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid")

            # F Row
            ws.cell(row=r_f, column=8, value="F").font = font_red_f
            ws.cell(row=r_f, column=8).alignment = align_center
            hrs_f = row.get("ชั่วโมงสอบ_F", "")
            ws.cell(row=r_f, column=9, value=hrs_f).alignment = align_center
            ws.cell(row=r_f, column=10, value=row.get("วันเวลาสอบ_F", "")).alignment = align_center
            ws.cell(row=r_f, column=11, value=row.get("ผู้คุมสอบ_F", "")).alignment = align_center

            if not hrs_f or hrs_f == "":
                ws.cell(row=r_f, column=9).fill = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid")

            for r in [r_m, r_f]:
                for c in range(1, 12):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border_all
                    if not cell.font or cell.font.name != "TH SarabunPSK":
                        cell.font = font_data

            global_item_no += 1
            current_row += 2

    # ปรับขนาดคอลัมน์
    col_widths = {"A": 5, "B": 12, "C": 28, "D": 18, "E": 14, "F": 10, "G": 6, "H": 6, "I": 8, "J": 32, "K": 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def create_dummy_subject_excel():
    data = {
        "คณะ": [
            "คณะเทคโนโลยีอุตสาหกรรมการเกษตร",
            "คณะเทคโนโลยีอุตสาหกรรมการเกษตร",
            "คณะเทคโนโลยีสังคม",
            "คณะเทคโนโลยีสังคม",
        ],
        "รหัสวิชา": ["05-110-104", "05-110-104", "02-303-101", "02-303-102"],
        "ชื่อวิชา": [
            "ภาษาอังกฤษทั่วไป",
            "ภาษาอังกฤษทั่วไป",
            "การวิเคราะห์และออกแบบระบบ",
            "การเขียนโปรแกรมคอมพิวเตอร์",
        ],
        "กลุ่มเรียน": ["IS-261", "BA-21", "IS-261", "IT-11"],
        "จำนวนผู้เข้าสอบ": [35, 38, 35, 40],
        "ชื่อผู้สอน": [
            "อ.มาริสา คงดี",
            "อ.สมชาย ใจดี",
            "ดร.สมชาย ใจดี",
            "อ.สมศรี รักเรียน",
        ],
        "สังกัดสาขา": [
            "ภาษาศาสตร์",
            "ภาษาศาสตร์",
            "เทคโนโลยีสารสนเทศ",
            "เทคโนโลยีสารสนเทศ",
        ],
        "ประเภทการสอบ": ["ทฤษฎี", "ทฤษฎี", "ทฤษฎี", "ปฏิบัติคอมพิวเตอร์"],
        "ชั่วโมงสอบ_M": [1.5, 1.5, 2, 2],
        "ชั่วโมงสอบ_F": [2, 2, 3, 2],
    }
    df = pd.DataFrame(data)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Exam_Subjects")
    return buffer.getvalue()


# ==================== 5. หน้าจอหลักและการทำงาน (Main App UI) ====================
if not st.session_state["logged_in"]:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.title("ระบบจัดตารางสอบอัตโนมัติ")
        st.subheader("งานส่งเสริมวิชาการและงานทะเบียน วิทยาเขตจันทบุรี")

        username_input = st.text_input("ชื่อผู้ใช้งาน (Username)")
        password_input = st.text_input("รหัสผ่าน (Password)", type="password")

        if st.button("เข้าสู่ระบบ 🔐", use_container_width=True):
            if (
                username_input in USER_CREDENTIALS
                and USER_CREDENTIALS[username_input] == password_input
            ):
                st.session_state["logged_in"] = True
                st.session_state["username"] = username_input
                st.rerun()
            else:
                st.error("ชื่อผู้ใช้งานหรือรหัสผ่านไม่ถูกต้อง")

else:
    header_col1, header_col2 = st.columns([8, 2])
    with header_col1:
        st.title("🏫 ระบบจัดการตารางสอบ")
        st.caption("มหาวิทยาลัยเทคโนโลยีราชมงคลตะวันออก วิทยาเขตจันทบุรี")
    with header_col2:
        st.write(f"ผู้ใช้งาน: **{st.session_state['username']}**")
        if st.button("ออกจากระบบ 🚪", on_click=logout, use_container_width=True):
            pass

    st.markdown("---")

    # --- SideBar ---
    st.sidebar.header("เมนูการใช้งาน")
    menu_selection = st.sidebar.radio(
        "เลือกฟังก์ชันที่ต้องการทำงาน:",
        [
            "1. จัดตารางสอบ (สะสมในระบบ)",
            "2. จัดการข้อมูลห้องสอบ",
        ],
    )

    st.sidebar.markdown("---")
    st.sidebar.header("🗓️ กำหนดช่วงเวลาสอบประจำเทอม")

    st.sidebar.subheader("1. สอบกลางภาค (Midterm)")
    m_start = st.sidebar.date_input("วันเริ่มสอบกลางภาค", datetime.date(2026, 8, 24))
    m_end = st.sidebar.date_input("วันสิ้นสุดสอบกลางภาค", datetime.date(2026, 8, 28))

    st.sidebar.subheader("2. สอบปลายภาค (Final)")
    f_start = st.sidebar.date_input("วันเริ่มสอบปลายภาค", datetime.date(2026, 10, 26))
    f_end = st.sidebar.date_input("วันสิ้นสุดสอบปลายภาค", datetime.date(2026, 11, 1))

    daily_slots = ["09:00 - 11:00", "11:00 - 13:00", "13:30 - 15:30", "15:30 - 17:30"]

    st.sidebar.markdown("---")
    st.sidebar.subheader("⚙️ จัดการข้อมูลสะสม")
    if st.sidebar.button("🔴 ล้างข้อมูลตารางสอบเดิม (เริ่มเทอมใหม่)", use_container_width=True):
        reset_semester_data()

    st.sidebar.download_button(
        label="ดาวน์โหลดไฟล์ตัวอย่างวิชาสอบ (.xlsx)",
        data=create_dummy_subject_excel(),
        file_name="exam_subjects_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ฐานข้อมูลห้องสอบ (Mock Data)
    df_rooms = pd.DataFrame([
        {"อาคาร": "อาคารเรียนรวม 36 ปี", "รหัสห้อง": "36-301", "ความจุสอบ": 40, "ประเภท": "ห้องทฤษฎี", "สถานะ": "ใช้งานได้"},
        {"อาคาร": "อาคารเรียนรวม 36 ปี", "รหัสห้อง": "36-302", "ความจุสอบ": 40, "ประเภท": "ห้องทฤษฎี", "สถานะ": "ใช้งานได้"},
        {"อาคาร": "อาคารปฏิบัติการไอที", "รหัสห้อง": "IT-201", "ความจุสอบ": 35, "ประเภท": "ห้องปฏิบัติการคอมพิวเตอร์", "สถานะ": "ใช้งานได้"},
    ])

    # ---------------- เมนูที่ 1: จัดตารางสอบ ----------------
    if menu_selection == "1. จัดตารางสอบ (สะสมในระบบ)":
        st.header("🗓️ จัดตารางสอบอัตโนมัติ (รองรับหลายคณะ & เพิ่มเติมวิชาสอบ)")
        
        # แสดงสถานะตารางสอบเดิมที่มีในระบบ
        history_df = st.session_state["history_schedule"]
        if not history_df.empty:
            st.info(f"💡 **ขณะนี้มีวิชาที่จัดสอบไปแล้วในเทอมนี้จำนวน {len(history_df)} รายวิชา** (การจัดสอบเพิ่มเติมรอบนี้จะเช็คไม่ให้เวลา/ห้อง/อาจารย์ ชนกับ {len(history_df)} วิชาเดิม)")
            with st.expander("🔍 ดูรายการตารางสอบเดิมที่มีในระบบ"):
                st.dataframe(history_df)
        else:
            st.write("📌 ยังไม่มีตารางสอบสะสมในเทอมนี้ (เป็นการจัดสอบรอบแรก)")

        uploaded_file = st.file_uploader(
            "เลือกไฟล์ข้อมูลวิชาสอบที่ต้องการเพิ่ม/จัดสอบ (.xlsx)", type=["xlsx"], key="main_uploader"
        )

        shuffle_invig = st.checkbox("ต้องการคละอาจารย์ผู้คุมสอบ (ถ้าไม่ติ๊ก ระบบจะให้อาจารย์ผู้สอนเป็นผู้คุมสอบ)", value=False)

        if uploaded_file is not None:
            df_uploaded = pd.read_excel(uploaded_file)
            st.write("📋 **รายการวิชาสอบใหม่ที่นำเข้า:**")
            st.dataframe(df_uploaded)

            if st.button("เริ่มประมวลผลจัดตารางสอบ ⚡", type="primary"):
                with st.spinner("กำลังคำนวณและตรวจสอบเงื่อนไขเวลา..."):
                    slots_m = generate_time_slots(m_start, m_end, daily_slots)
                    slots_f = generate_time_slots(f_start, f_end, daily_slots)

                    # คำนวณตารางสอบวิชาใหม่ โดยส่งตารางสอบเดิมไปเช็คชน
                    df_new_result = auto_schedule_exams(
                        df_uploaded,
                        df_rooms,
                        slots_m,
                        slots_f,
                        existing_schedule=history_df,
                        shuffle_invigilators=shuffle_invig,
                    )

                    # รวมผลลัพธ์วิชาใหม่เข้ากับประวัติเดิม
                    updated_schedule = pd.concat([history_df, df_new_result], ignore_index=True)
                    st.session_state["history_schedule"] = updated_schedule

                st.balloons()
                st.success(f"✅ จัดตารางสอบสำเร็จ! เพิ่มวิชาสอบใหม่เรียบร้อยแล้ว (รวมทั้งหมดในเทอมนี้ {len(updated_schedule)} รายวิชา)")
                
                st.subheader("📊 ตารางสอบรวมทั้งหมดในระบบปัจจุบัน")
                st.dataframe(updated_schedule)

                str_m = f"{m_start.strftime('%d/%m/%Y')} - {m_end.strftime('%d/%m/%Y')}"
                str_f = f"{f_start.strftime('%d/%m/%Y')} - {f_end.strftime('%d/%m/%Y')}"
                excel_data = generate_excel_report(updated_schedule, str_m, str_f)

                st.download_button(
                    label="ดาวน์โหลดตารางสอบรวม Excel (รูปแบบ MF) 📥",
                    data=excel_data,
                    file_name="ตารางสอบรวม_1_2569_MF.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ---------------- เมนูที่ 2: จัดการข้อมูลห้องสอบ ----------------
    elif menu_selection == "2. จัดการข้อมูลห้องสอบ":
        st.header("🏫 จัดการข้อมูลห้องสอบ")
        st.table(df_rooms)

    
